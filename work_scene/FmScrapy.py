# -*- coding: utf-8 -*-
"""
@Name:        fund_management_fetch.py
@Describe:    XXX
@Time:        2025/01/16 00:38:38
@Author:      Ray
@Version:     1.0
"""
import re
import json
import time
import requests
import openpyxl
import asyncio
import aiohttp
import pandas as pd
from bs4 import BeautifulSoup
from googletrans import Translator
from openpyxl.utils import get_column_letter


def gene_txt_from_xml(xml):
    soup = BeautifulSoup(xml, "xml")
    # 提取纯文本
    plain_text = soup.get_text()
    return plain_text


def generate_module_info(file):
    # 获取result.json文件，包含所有FM功能模块
    url = "https://fioriappslibrary.hana.ondemand.com/sap/fix/externalViewer/services/SingleApp.xsodata/$batch"
    payload = """
		\r\n--batch_668f-7011-a43b\r\nContent-Type: application/http\r\nContent-Transfer-Encoding: binary\r\n\r\nGET Details(inpfioriId='W0189',inpreleaseId='S26OP',inpLanguage='None',fioriId='W0189',releaseId='S26OP')/RolesPerApp(InpFioriId='W0189',fioriId='W0189',releaseId='S26OP',RoleName='Budget%20Responsible%20-%20Funds%20Management',releaseGroupText='SAP%20S%2F4HANA%20(Private%20Cloud%20and%20On-Premise)')/RelatedAppsByRole?$skip=0&$top=10000&$orderby=AppName%20asc&$inlinecount=allpages HTTP/1.1\r\nAccept-Language: en\r\nAccept: application/json\r\nMaxDataServiceVersion: 2.0\r\nDataServiceVersion: 2.0\r\n\r\n\r\n--batch_668f-7011-a43b--\r\n
	"""
    headers = {'Content-Type': 'multipart/mixed;boundary=batch_668f-7011-a43b'}
    response = requests.request("POST", url, headers=headers, data=payload)
    # print(response.status_code)
    if response.status_code == 202:
        result = re.findall(r"({.*})", response.text)
        with open(file, "w", encoding="utf-8") as file:
            json_data = json.loads(result[0])
            json.dump(json_data, file, indent=4, ensure_ascii=False)
        print("FM模块信息写入result.json文件")
    else:
        print("FM模块信息获取失败！！")


def generate_xlsx_file(file):
    # 生成xlsx文件
    with open(file, 'r') as file:
        data = json.load(file)
    fioriid_appname = [{'fioriId': 'W0189',
                        'AppName': 'Actual Line Item Analyzer',
                        'APPInfo': None,
                        'StartVersion': None}]
    # 这里假设你的JSON数据是一个列表，里面包含多个字典
    selected_item = {}
    for item in data['d']['results']:
        selected_item['fioriId'] = item.get('fioriId')
        selected_item['AppName'] = item.get('AppName')
        selected_item['APPInfo'] = None
        selected_item['StartVersion'] = None
        fioriid_appname.append(selected_item)
        selected_item = {}
    release_id = {
        'S15OP': 'S / 4HANA 1909',
        'S16OP': 'S / 4HANA 1909 FPS01',
        'S17OP': 'S / 4HANA 1909 FPS02',
        'S18OP': 'S / 4HANA 2020',
        'S19OP': 'S / 4HANA 2020 FPS01',
        'S20OP': 'S / 4HANA 2020 FPS02',
        'S21OP': 'S / 4HANA 2021',
        'S22OP': 'S / 4HANA 2021 FPS01',
        'S23OP': 'S / 4HANA 2021 FPS02',
        'S24OP': 'S / 4HANA 2022',
        'S25OP': 'S / 4HANA 2022 FPS01',
        'S26OP': 'S / 4HANA 2022 FPS02',
        'S27OP': 'S / 4HANA 2023',
        'S28OP': 'S / 4HANA 2023 FPS01',
        'S29OP': 'S / 4HANA 2023 FPS02',
        'S29PCE': 'S / 4HANA 2023 FPS02 - PCE'}
    tmp_data = []
    count = 0
    for item in fioriid_appname:
        count += 1
        tmp_dict = item
        pre_url = "https://fioriappslibrary.hana.ondemand.com/sap/fix/externalViewer/services/getGTMDescription.xsjs?fioriId={}".format(
            item['fioriId'])
        sample = ""
        flag = True
        for key, value in release_id.items():
            url = pre_url + '&releaseId=' + key
            payload = {}
            headers = {
                'Cookie': 'sapxslb=BBBE8BA0B5C1C7409C1DA71C628CB438; BIGipServerboma0d717969.factory.customdomain=!x5bNrMOYQ8/cJbheaBxHSKFoIXUAnB75Dnb/+QwDUWw9oUBqdPs0vW4ljTzOSU22yinSrCrNUDeDw6o=; language=None; xsSecureId5ABB8FCC7F4EC1526A2A6CE5FFD964CF=8C83D5235FB15245A563D8ADA2ADF2F4',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'cross-site'}
            response = requests.request(
                "GET", url, headers=headers, data=payload)
            json_data = response.json()
            app_key_features = json_data["AppKeyFeatures"]
            feature_description = json_data["GTMAppDescription"]
            if flag and feature_description != "":
                tmp_dict["APPInfo"] = gene_txt_from_xml(feature_description)
                tmp_dict["StartVersion"] = release_id[key]
                flag = False
            if app_key_features == sample and sample != "":
                tmp_dict[value] = ""
                continue
            else:
                tmp_dict[value] = gene_txt_from_xml(app_key_features)
                sample = app_key_features
        tmp_data.append(tmp_dict)

        print("已经处理完第: {}条记录".format(count))
        # if count > 5:
        # 	break
    # 将选定的数据转换为DataFrame
    df = pd.DataFrame(tmp_data)
    # 写入Excel文件
    df.to_excel('result_1.xlsx', index=False, engine='openpyxl')
    print("\n数据已成功写入Excel文件！")


async def translate_text(text):
    translator = Translator()
    translated = await translator.translate(text, src="en", dest="ja")
    return translated.text

# Run the async function


def translate_result_file():
    input_file = "./work_scene/outputs/output.xlsx"
    output_file = "./work_scene/outputs/output_jp.xlsx"
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active  # 获取活动工作表
    # 指定需要翻译的列（可以按列名或索引）
    columns_not_translate = ["A", "B", "D"]  # 指定需要翻译的列（按列号，如 A 列和 C 列）
    max_column = sheet.max_column
    # 获取所有列标题（A, B, C, ...）
    columns_translate = [get_column_letter(col) for col in range(
        1, max_column + 1) if get_column_letter(col) not in columns_not_translate]
    # 初始化翻译器
    # translator = Translator()

    # 遍历指定列并翻译内容
    for row in range(2, sheet.max_row + 1):  # 从第2行开始，跳过标题行
        for col in columns_translate:
            cell = sheet[f"{col}{row}"]  # 获取指定列单元格
            if cell.value:  # 如果单元格有值
                try:
                    cell.value = asyncio.run(
                        translate_text(cell.value))  # 异步翻译
                except Exception as e:
                    print(f"翻译失败: {cell.value} -> {e}")
        print("已经翻译第: {}条记录".format(row - 1))
    # 保存翻译后的 Excel 文件
    wb.save(output_file)
    print(f"翻译完成，结果已保存到 {output_file}")


def fetch_url():
    release_id = {
        'S15OP': 'S / 4HANA 1909',
        'S16OP': 'S / 4HANA 1909 FPS01',
        'S17OP': 'S / 4HANA 1909 FPS02',
        'S18OP': 'S / 4HANA 2020',
        'S19OP': 'S / 4HANA 2020 FPS01',
        'S20OP': 'S / 4HANA 2020 FPS02',
        'S21OP': 'S / 4HANA 2021',
        'S22OP': 'S / 4HANA 2021 FPS01',
        'S23OP': 'S / 4HANA 2021 FPS02',
        'S24OP': 'S / 4HANA 2022',
        'S25OP': 'S / 4HANA 2022 FPS01',
        'S26OP': 'S / 4HANA 2022 FPS02',
        'S27OP': 'S / 4HANA 2023',
        'S28OP': 'S / 4HANA 2023 FPS01',
        'S29OP': 'S / 4HANA 2023 FPS02',
        'S29PCE': 'S / 4HANA 2023 FPS02 - PCE'}
    # 生成xlsx文件
    with open('./work_scene/inputs/result.json', 'r') as file:
        data = json.load(file)
    # 这里假设你的JSON数据是一个列表，里面包含多个字典
    foriids = [item.get('fioriId') for item in data['d']['results']]
    appnames = [item.get('AppName') for item in data['d']['results']]
    foriids.insert(0, 'W0189')
    appnames.insert(0, 'Actual Line Item Analyzer')
    versions = release_id.keys()
    foriids_urls = {}

    for foriid in foriids:
        urls = []
        for version in versions:
            url = "https://fioriappslibrary.hana.ondemand.com/sap/fix/externalViewer/services/getGTMDescription.xsjs?fioriId={}&releaseId={}".format(
                foriid, version)
            urls.append(url)
        foriids_urls[foriid] = urls
    foriid_appname = dict(zip(foriids, appnames))
    return foriid_appname, foriids_urls


async def fetch(session, url):
    headers = {
        'Cookie': 'sapxslb=BBBE8BA0B5C1C7409C1DA71C628CB438; BIGipServerboma0d717969.factory.customdomain=!x5bNrMOYQ8/cJbheaBxHSKFoIXUAnB75Dnb/+QwDUWw9oUBqdPs0vW4ljTzOSU22yinSrCrNUDeDw6o=; language=None; xsSecureId5ABB8FCC7F4EC1526A2A6CE5FFD964CF=8C83D5235FB15245A563D8ADA2ADF2F4',
        'Sec-Fetch-Dest': 'document', 'Sec-Fetch-Mode': 'navigate', 'Sec-Fetch-Site': 'cross-site'}
    async with session.get(url, headers=headers, ssl=False) as response:
        return await response.json()


async def async_fetchs(fioriid, appname, urls):
    results = []
    async with aiohttp.ClientSession(trust_env=True) as session:
        tasks = [fetch(session, url) for url in urls]
        results = await asyncio.gather(*tasks)
    # fioriid_appname = [{'fioriId': 'W0189', 'AppName': 'Actual Line Item Analyzer', 'APPInfo': None, 'StartVersion': None}]
    release_id = {'S15OP': 'S / 4HANA 1909', 'S16OP': 'S / 4HANA 1909 FPS01', 'S17OP': 'S / 4HANA 1909 FPS02',
                  'S18OP': 'S / 4HANA 2020', 'S19OP': 'S / 4HANA 2020 FPS01', 'S20OP': 'S / 4HANA 2020 FPS02',
                  'S21OP': 'S / 4HANA 2021', 'S22OP': 'S / 4HANA 2021 FPS01', 'S23OP': 'S / 4HANA 2021 FPS02',
                  'S24OP': 'S / 4HANA 2022', 'S25OP': 'S / 4HANA 2022 FPS01', 'S26OP': 'S / 4HANA 2022 FPS02',
                  'S27OP': 'S / 4HANA 2023', 'S28OP': 'S / 4HANA 2023 FPS01', 'S29OP': 'S / 4HANA 2023 FPS02',
                  'S29PCE': 'S / 4HANA 2023 FPS02 - PCE'
                  }
    versions = [url.split("=")[-1] for url in urls]
    # print(versions)
    selected_item = dict()
    selected_item["fioriid"] = fioriid
    selected_item["AppName"] = appname
    selected_item["AppInfo"] = None
    selected_item["StartVersion"] = None
    flag = True
    for index in range(len(results)):
        if flag and results[index]["GTMAppDescription"] != "":
            selected_item["StartVersion"] = release_id[versions[index]]
            # print(results[index]["GTMAppDescription"])
            selected_item["AppInfo"] = gene_txt_from_xml(
                results[index]["GTMAppDescription"])
            flag = False
        if results[index]["AppKeyFeatures"] != "":
            sample = gene_txt_from_xml(results[index]["AppKeyFeatures"])
            if sample in selected_item.values():
                selected_item[release_id[versions[index]]] = ''
            else:
                selected_item[release_id[versions[index]]] = sample
        else:
            selected_item[release_id[versions[index]]] = ''

    return selected_item


if __name__ == "__main__":
    module_info_file = "./work_scene/inputs/result.json"
    # generate_module_info(module_info_file)

    # 同步处理，效率低，大概28分钟
    # generate_xlsx_file(module_info_file)

    # 异步处理，效率高，1.5分钟
    start_time = time.time()
    fioriid_appname, urls = fetch_url()
    fioriid_appname_urls = []
    fioriids = list(fioriid_appname.keys())
    for fioriId, url in urls.items():
        fioriid_appname_url = (fioriId, fioriid_appname[fioriId], url)
        fioriid_appname_urls.append(fioriid_appname_url)
    selected_items = []
    for item in fioriid_appname_urls:
        fioriid = item[0]
        appname = item[1]
        url = item[2]
        selected_item = asyncio.run(async_fetchs(fioriid, appname, url))
        selected_items.append(selected_item)
        print("已经处理完第: {}条记录".format(fioriids.index(fioriid) + 1))
    # 写入到 Excel 文件
    df = pd.DataFrame(selected_items)
    df.to_excel('output.xlsx', index=False)
    print("数据已写入 output.xlsx 文件！")
    end_time = time.time()
    print(end_time - start_time)

    # 翻译成日语
    # translate_result_file()
